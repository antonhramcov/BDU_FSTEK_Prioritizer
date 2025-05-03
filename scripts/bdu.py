from openpyxl import load_workbook
from string import ascii_uppercase
from typing import Union
from re import match, findall
import json

def check_id_bdu(id:str) -> bool:
    return match(r'BDU:\d{4}-\d+', id)

def check_id_cve(id:str) -> bool:
    return match(r'CVE-\d{4}-\d+', id)

def get_dict_from_xlsx():
    sheet = load_workbook('../vullist.xlsx').active
    return {sheet[f"A{row}"].value : str(sheet[f"S{row}"].value) for row in range(4, sheet.max_row+1)}

def from_xlsx_to_json():
    with open('../database.json', 'w', encoding='utf-8') as f:
        json.dump(get_dict_from_xlsx(), f, ensure_ascii=False, indent=4)

def search_id(id:str) -> Union[str, None]:
    if check_id_bdu(id):
        try:
            with open('../database.json', 'r', encoding='utf-8') as f:
                data = json.load(f)
                search = data.get(id)
                if search:
                    if check_id_cve(search):
                        return findall(r'CVE-\d{4}-\d+', search)[0]
                    else:
                        return search
                else:
                    return f'BDU_ID {id} не найден '

        except FileNotFoundError:
            return 'Отсутствует файл database.json'
    else:
        return f'Неправильный BDU_ID - {id}'
