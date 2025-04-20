from openpyxl import load_workbook
from string import ascii_uppercase
from typing import Union
from re import match, findall


def find(id:str) -> Union[dict, str, None]:
    sheet = load_workbook('vullist.xlsx').active
    if not match(r'BDU:\d{4}-\d+', id):
        return f'{id} Error: BDUs should be provided in the standard format BDU:0000-00000'
    for row in range(4,sheet.max_row+1):
        if id == sheet[f"A{row}"].value:
            values = [sheet[f"{x}{row}"].value for x in ascii_uppercase[:-1]]
            keys = [sheet[f"{x}3"].value for x in ascii_uppercase[:-1]]
            return dict(zip(keys, values))
    return None

def get_level(id:str) -> str:
    if find(id) != None:
        value_level = find(id)['Уровень опасности уязвимости'].split()[0]
        name = {'Низкий':'Priority 4', 'Средний':'Priority 3', 'Высокий':'Priority 2', 'Критический':'Priority 1'}
        return name[value_level]


def bdu_from_cve(id:str) -> Union[str, None]:
    sheet = load_workbook('vullist.xlsx').active
    if not match(r'BDU:\d{4}-\d+', id):
        return f'{id} Error: BDUs should be provided in the standard format BDU:0000-00000'
    for row in range(4, sheet.max_row+1):
        if id == sheet[f"A{row}"].value:
            value = sheet[f"S{row}"].value
            if match(r'(CVE-\d{4}-\d+)', value):
                return findall(r'(CVE-\d{4}-\d+)', value)[0]
            else:
                return None
